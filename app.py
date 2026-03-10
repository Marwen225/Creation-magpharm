import streamlit as st
import pandas as pd
from io import BytesIO
import os
import shutil
import openpyxl
from copy import copy

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Reference files (read-only, shipped with the repo)
MEDECINS_TEMPLATE = os.path.join(BASE_DIR, "Médecins.xlsx")
PHARMACIES_TEMPLATE = os.path.join(BASE_DIR, "Pharmacies.xlsx")
ONEKEY_MED_FILE = os.path.join(BASE_DIR, "One key medecin.xlsx")
ONEKEY_PHA_FILE = os.path.join(BASE_DIR, "one key pharmacie.xlsx")
UTILISATEURS_FILE = os.path.join(BASE_DIR, "Utilisateurs.xlsx")

# Writable data directory (Railway volume or local fallback)
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "data"))
os.makedirs(DATA_DIR, exist_ok=True)

# Writable Excel files (auto-copied from templates on first run)
MEDECINS_FILE = os.path.join(DATA_DIR, "Médecins.xlsx")
PHARMACIES_FILE = os.path.join(DATA_DIR, "Pharmacies.xlsx")

if not os.path.exists(MEDECINS_FILE):
    shutil.copy2(MEDECINS_TEMPLATE, MEDECINS_FILE)
if not os.path.exists(PHARMACIES_FILE):
    shutil.copy2(PHARMACIES_TEMPLATE, PHARMACIES_FILE)

DOCTOR_COLUMNS = [
    "name", "ref", "type_id", "email", "phone", "mobile",
    "customer_potential_id/id", "doctor_speciality_id",
    "doctor_speciality2_id", "doctor_status", "doctor_grade_id",
    "doctor_institution_id", "department_id", "street", "street2",
    "country_id", "state_id", "Commune", "city_id/id", "zip",
    "Secteur (nom complet)", "sector_id/id",
    "property_product_pricelist", "static_portfolio_user_ids",
]

PHARMACY_COLUMNS = [
    "name", "ref", "type_id", "email", "phone", "mobile",
    "customer_potential_id/id", "street", "street2", "country_id",
    "state_id", "Commune", "city_id/id", "zip",
    "Secteur (nom complet)", "sector_id/id",
    "company_registry", "vat", "tax_article", "sin",
    "property_product_pricelist", "static_portfolio_user_ids",
]


# ---------------------------------------------------------------------------
# Load reference data — read raw column headers from Excel
# ---------------------------------------------------------------------------
@st.cache_data
def load_adresses():
    """Load Adresses sheet.
    Excel layout:
      A=state  B=(empty)  C=id  D=name(COMMUNE)  E=state_id(WILAYA)  F=(empty)  G=sector  H=id(sector_id)
    """
    df = pd.read_excel(MEDECINS_TEMPLATE, sheet_name="Adresses", header=0)

    # Communes: column D='name' → commune, column E='state_id' → wilaya, column C='id' → city_id
    communes_df = df[["name", "state_id", "id"]].copy()
    communes_df.columns = ["commune", "wilaya", "city_id"]
    communes_df = communes_df.dropna(subset=["commune", "wilaya"]).drop_duplicates(
        subset=["commune", "wilaya"]
    )

    # Sectors: column G='sector', column H='id.1' → sector_id
    sectors_df = df[["sector", "id.1"]].copy()
    sectors_df.columns = ["sector", "sector_id"]
    sectors_df = (
        sectors_df.dropna(subset=["sector"])
        .drop_duplicates(subset=["sector"])
        .sort_values("sector")
    )

    wilayas = sorted(communes_df["wilaya"].unique().tolist())
    return wilayas, communes_df, sectors_df


@st.cache_data
def load_medical():
    df = pd.read_excel(MEDECINS_TEMPLATE, sheet_name="Médical", header=0)
    specialities = sorted(df["speciality"].dropna().unique().tolist())
    institutions = sorted(df["institution"].dropna().unique().tolist())
    grades = sorted(df["grade"].dropna().unique().tolist())
    departments = sorted(df["department"].dropna().unique().tolist())
    statuses = sorted(df["status"].dropna().unique().tolist())
    return specialities, institutions, grades, departments, statuses


@st.cache_data
def load_legendes_med():
    df = pd.read_excel(MEDECINS_TEMPLATE, sheet_name="Legendes", header=0)
    potentials = df.iloc[:, 0].dropna().unique().tolist()
    pricelists = df.iloc[:, 2].dropna().unique().tolist()
    types = df.iloc[:, 4].dropna().unique().tolist()
    return potentials, pricelists, types


@st.cache_data
def load_legendes_pha():
    df = pd.read_excel(PHARMACIES_TEMPLATE, sheet_name="Legendes", header=0)
    potentials = df.iloc[:, 0].dropna().unique().tolist()
    pricelists = df.iloc[:, 2].dropna().unique().tolist()
    types = df.iloc[:, 4].dropna().unique().tolist()
    return potentials, pricelists, types


@st.cache_data
def load_onekey_medecins():
    """Load doctor names (and specialties) from One Key file."""
    df = pd.read_excel(ONEKEY_MED_FILE, usecols=["Account: Account Name", "Account: Specialty"])
    df = df.dropna(subset=["Account: Account Name"])
    df["Account: Account Name"] = df["Account: Account Name"].astype(str).str.strip()
    return df


@st.cache_data
def load_onekey_pharmacies():
    """Load pharmacy names from One Key file."""
    df = pd.read_excel(ONEKEY_PHA_FILE, usecols=["Account: Account Name"])
    df = df.dropna(subset=["Account: Account Name"])
    names = df["Account: Account Name"].astype(str).str.strip().unique().tolist()
    return sorted(names)


@st.cache_data
def load_utilisateurs():
    """Load delegate names from Utilisateurs.xlsx."""
    df = pd.read_excel(UTILISATEURS_FILE, usecols=["Nom"])
    df = df.dropna(subset=["Nom"])
    names = df["Nom"].astype(str).str.strip().unique().tolist()
    return sorted(names)


# ---------------------------------------------------------------------------
# Lookups
# ---------------------------------------------------------------------------
def get_city_id(communes_df, commune, wilaya):
    m = communes_df[(communes_df["commune"] == commune) & (communes_df["wilaya"] == wilaya)]
    return m.iloc[0]["city_id"] if not m.empty else ""


def get_sector_id(sectors_df, sector):
    m = sectors_df[sectors_df["sector"] == sector]
    return m.iloc[0]["sector_id"] if not m.empty else ""


def get_communes_for_wilaya(communes_df, wilaya):
    if not wilaya:
        return []
    return communes_df.loc[communes_df["wilaya"] == wilaya, "commune"].sort_values().tolist()


# ---------------------------------------------------------------------------
# Write directly into existing Excel files
# ---------------------------------------------------------------------------
def append_to_excel(filepath, sheet_name, columns, df):
    """Append rows from df into the specified sheet of the existing Excel file.
    Writes values starting at the first empty row after the header."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    # Build header→column index map from row 1
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col_idx).value
        if h:
            header_map[h] = col_idx

    # Find first empty row
    start_row = 2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            start_row = r
            break

    # Write each record
    for i, (_, record) in enumerate(df.iterrows()):
        row_num = start_row + i
        for col_name in columns:
            if col_name in header_map and col_name in record.index:
                val = record[col_name]
                if pd.isna(val):
                    val = None
                ws.cell(row=row_num, column=header_map[col_name], value=val)

    wb.save(filepath)
    return start_row, start_row + len(df) - 1


def load_existing_records(filepath, sheet_name):
    """Load existing records from the Excel file for duplicate detection."""
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)
        return df
    except Exception:
        return pd.DataFrame()


def generate_excel_download(df, columns, sheet_name):
    """Generate a standalone Excel download (backup option)."""
    export_df = pd.DataFrame(columns=columns)
    for col in columns:
        if col in df.columns:
            export_df[col] = df[col].values
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
def init_state():
    if "doctors" not in st.session_state:
        st.session_state.doctors = pd.DataFrame(columns=DOCTOR_COLUMNS)
    if "pharmacies" not in st.session_state:
        st.session_state.pharmacies = pd.DataFrame(columns=PHARMACY_COLUMNS)


# ---------------------------------------------------------------------------
# Doctor form
# ---------------------------------------------------------------------------
def doctor_form():
    st.header("Créer un Médecin")

    wilayas, communes_df, sectors_df = load_adresses()
    specialities, institutions, grades, departments, statuses = load_medical()
    potentials, pricelists, types_med = load_legendes_med()
    onekey_med_df = load_onekey_medecins()

    # --- Name search with suggestions (outside form for instant refresh) ---
    st.subheader("🔍 Nom du médecin")
    search_name = st.text_input("Rechercher un nom (min 3 caractères) *", key="d_name_search")
    selected_name = ""
    onekey_specialty = ""
    if len(search_name.strip()) >= 3:
        mask = onekey_med_df["Account: Account Name"].str.contains(
            search_name.strip(), case=False, na=False
        )
        matches = onekey_med_df.loc[mask].head(50)
        if not matches.empty:
            options = matches["Account: Account Name"].tolist()
            choice = st.selectbox(
                f"Suggestions ({len(matches)} résultats, max 50 affichés)",
                ["(saisie libre)"] + options,
                key="d_name_select",
            )
            if choice != "(saisie libre)":
                selected_name = choice
                # Auto-fill specialty from One Key
                spec_row = matches[matches["Account: Account Name"] == choice]
                if not spec_row.empty and pd.notna(spec_row.iloc[0]["Account: Specialty"]):
                    onekey_specialty = str(spec_row.iloc[0]["Account: Specialty"]).strip()
            else:
                selected_name = search_name.strip()
        else:
            st.info("Aucun résultat trouvé dans One Key. Le nom saisi sera utilisé.")
            selected_name = search_name.strip()
    elif search_name.strip():
        st.caption("Tapez au moins 3 caractères pour voir les suggestions.")
        selected_name = search_name.strip()

    # --- Wilaya → Commune (outside form for instant refresh) ---
    st.subheader("📍 Localisation")
    c1, c2, c3 = st.columns(3)
    with c1:
        wilaya = st.selectbox("Wilaya *", [""] + wilayas, key="d_wil")
    with c2:
        commune_list = get_communes_for_wilaya(communes_df, wilaya)
        commune = st.selectbox(
            f"Commune * ({len(commune_list)} disponibles)",
            [""] + commune_list,
            key="d_com",
        )
    with c3:
        sector = st.selectbox("Secteur *", [""] + sectors_df["sector"].tolist(), key="d_sec")

    # --- Delegate search (outside form for instant refresh) ---
    st.subheader("Délégué")
    utilisateurs = load_utilisateurs()
    d_del_search = st.text_input("Rechercher votre nom (min 3 caractères) *", key="d_del_search")
    selected_delegate = ""
    if len(d_del_search.strip()) >= 3:
        query_del = d_del_search.strip().lower()
        del_matches = [n for n in utilisateurs if query_del in n.lower()][:50]
        if del_matches:
            del_choice = st.selectbox(
                f"Sélectionner votre nom ({len(del_matches)} résultats)",
                [""] + del_matches,
                key="d_del_select",
            )
            selected_delegate = del_choice
        else:
            st.warning("Aucun résultat trouvé. Vérifiez l'orthographe.")
    elif d_del_search.strip():
        st.caption("Tapez au moins 3 caractères pour voir les suggestions.")

    # --- Other fields inside form ---
    with st.form("doc_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            type_id = st.selectbox("Type *", types_med)
            # Pre-select specialty from One Key if available
            spec_index = 0
            if onekey_specialty and onekey_specialty in specialities:
                spec_index = specialities.index(onekey_specialty) + 1
            speciality = st.selectbox("Spécialité *", [""] + specialities, index=spec_index)
            speciality2 = st.selectbox("Spécialité secondaire *", [""] + specialities)
            status = st.selectbox("Statut *", [""] + statuses)
            grade = st.selectbox("Grade *", [""] + grades)
            institution = st.selectbox("Institution *", [""] + institutions)
            department = st.selectbox("Département *", [""] + departments)
        with c2:
            street = st.text_input("Adresse *")
            phone = st.text_input("Téléphone *")
            mobile = st.text_input("Fax")
            email = st.text_input("Email")
            potential = st.selectbox("Potentiel *", [""] + potentials)
            pricelist = st.selectbox("Liste de prix", [""] + pricelists)

        submitted = st.form_submit_button("Ajouter le médecin")

    if submitted:
        name = selected_name
        errors = []
        if not name.strip():
            errors.append("Le nom est obligatoire.")
        if not speciality:
            errors.append("La spécialité est obligatoire.")
        if not speciality2:
            errors.append("La spécialité secondaire est obligatoire.")
        if not wilaya:
            errors.append("La wilaya est obligatoire.")
        if not commune:
            errors.append("La commune est obligatoire.")
        if not sector:
            errors.append("Le secteur est obligatoire.")
        if not status:
            errors.append("Le statut est obligatoire.")
        if not grade:
            errors.append("Le grade est obligatoire.")
        if not institution:
            errors.append("L'institution est obligatoire.")
        if not department:
            errors.append("Le département est obligatoire.")
        if not street.strip():
            errors.append("L'adresse est obligatoire.")
        if not phone.strip():
            errors.append("Le téléphone est obligatoire.")
        if not potential:
            errors.append("Le potentiel est obligatoire.")
        delegate = selected_delegate
        if not delegate:
            errors.append("Le délégué est obligatoire (sélectionnez depuis la liste).")

        # Check duplicates in session AND in existing Excel file
        if name.strip() and commune:
            dup = st.session_state.doctors[
                (st.session_state.doctors["name"].str.upper() == name.strip().upper())
                & (st.session_state.doctors["Commune"].str.upper() == commune.upper())
            ]
            if not dup.empty:
                errors.append("Ce médecin existe déjà dans la session (même nom + commune).")
            existing_xl = load_existing_records(MEDECINS_FILE, "Contacts")
            if not existing_xl.empty and "name" in existing_xl.columns and "Commune" in existing_xl.columns:
                dup_xl = existing_xl[
                    (existing_xl["name"].astype(str).str.upper() == name.strip().upper())
                    & (existing_xl["Commune"].astype(str).str.upper() == commune.upper())
                ]
                if not dup_xl.empty:
                    errors.append("Ce médecin existe déjà dans le fichier Excel (même nom + commune).")

        if errors:
            for e in errors:
                st.error(e)
        else:
            row = {
                "name": name.strip(),
                "ref": "",
                "type_id": type_id,
                "email": email.strip(),
                "phone": phone.strip(),
                "mobile": mobile.strip(),
                "customer_potential_id/id": potential or "",
                "doctor_speciality_id": speciality,
                "doctor_speciality2_id": speciality2 or "",
                "doctor_status": status,
                "doctor_grade_id": grade or "",
                "doctor_institution_id": institution or "",
                "department_id": department or "",
                "street": street.strip(),
                "street2": "",
                "country_id": "Algérie",
                "state_id": wilaya,
                "Commune": commune,
                "city_id/id": get_city_id(communes_df, commune, wilaya),
                "zip": "",
                "Secteur (nom complet)": sector,
                "sector_id/id": get_sector_id(sectors_df, sector),
                "property_product_pricelist": pricelist or "",
                "static_portfolio_user_ids": delegate,
            }
            row_df = pd.DataFrame([row])
            try:
                r_start, r_end = append_to_excel(
                    MEDECINS_FILE, "Contacts", DOCTOR_COLUMNS, row_df,
                )
                st.success(f"✅ Médecin **{row['name']}** enregistré dans Médecins.xlsx (ligne {r_start}) !")
            except Exception as exc:
                st.error(f"❌ Erreur d'écriture : {exc}")
            st.session_state.doctors = pd.concat(
                [st.session_state.doctors, row_df], ignore_index=True
            )

    # --- Preview ---
    if not st.session_state.doctors.empty:
        st.subheader(f"📋 Médecins ajoutés cette session ({len(st.session_state.doctors)})")
        st.dataframe(
            st.session_state.doctors[
                ["name", "doctor_speciality_id", "doctor_status",
                 "state_id", "Commune", "Secteur (nom complet)",
                 "phone", "static_portfolio_user_ids"]
            ],
            use_container_width=True, hide_index=True,
        )

        col_a, col_b = st.columns(2)
        with col_a:
            buf = generate_excel_download(st.session_state.doctors, DOCTOR_COLUMNS, "Contacts")
            st.download_button(
                "📥 Télécharger une copie", buf,
                file_name="Médecins_import.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col_b:
            if st.button("🗑️ Réinitialiser médecins"):
                st.session_state.doctors = pd.DataFrame(columns=DOCTOR_COLUMNS)
                st.rerun()


# ---------------------------------------------------------------------------
# Pharmacy form
# ---------------------------------------------------------------------------
def pharmacy_form():
    st.header("Créer une Pharmacie / Compte")

    wilayas, communes_df, sectors_df = load_adresses()
    potentials, pricelists, types_pha = load_legendes_pha()
    onekey_pha_names = load_onekey_pharmacies()

    # --- Name search with suggestions (outside form for instant refresh) ---
    st.subheader("🔍 Nom de la pharmacie")
    search_name = st.text_input("Rechercher un nom (min 3 caractères) *", key="p_name_search")
    selected_name = ""
    if len(search_name.strip()) >= 3:
        query = search_name.strip().lower()
        matches = [n for n in onekey_pha_names if query in n.lower()][:50]
        if matches:
            choice = st.selectbox(
                f"Suggestions ({len(matches)} résultats, max 50 affichés)",
                ["(saisie libre)"] + matches,
                key="p_name_select",
            )
            if choice != "(saisie libre)":
                selected_name = choice
            else:
                selected_name = search_name.strip()
        else:
            st.info("Aucun résultat trouvé dans One Key. Le nom saisi sera utilisé.")
            selected_name = search_name.strip()
    elif search_name.strip():
        st.caption("Tapez au moins 3 caractères pour voir les suggestions.")
        selected_name = search_name.strip()

    # --- Wilaya → Commune ---
    st.subheader("📍 Localisation")
    c1, c2, c3 = st.columns(3)
    with c1:
        wilaya = st.selectbox("Wilaya *", [""] + wilayas, key="p_wil")
    with c2:
        commune_list = get_communes_for_wilaya(communes_df, wilaya)
        commune = st.selectbox(
            f"Commune * ({len(commune_list)} disponibles)",
            [""] + commune_list,
            key="p_com",
        )
    with c3:
        sector = st.selectbox("Secteur *", [""] + sectors_df["sector"].tolist(), key="p_sec")

    # --- Delegate search (outside form for instant refresh) ---
    st.subheader("👤 Délégué")
    utilisateurs = load_utilisateurs()
    p_del_search = st.text_input("Rechercher votre nom (min 3 caractères) *", key="p_del_search")
    selected_delegate = ""
    if len(p_del_search.strip()) >= 3:
        query_del = p_del_search.strip().lower()
        del_matches = [n for n in utilisateurs if query_del in n.lower()][:50]
        if del_matches:
            del_choice = st.selectbox(
                f"Sélectionner votre nom ({len(del_matches)} résultats)",
                [""] + del_matches,
                key="p_del_select",
            )
            selected_delegate = del_choice
        else:
            st.warning("Aucun résultat trouvé. Vérifiez l'orthographe.")
    elif p_del_search.strip():
        st.caption("Tapez au moins 3 caractères pour voir les suggestions.")

    # --- Other fields ---
    with st.form("pha_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            type_id = st.selectbox("Type *", types_pha)
            street = st.text_input("Adresse *")
            phone = st.text_input("Téléphone *")
            mobile = st.text_input("Fax")
            email = st.text_input("Email")
            potential = st.selectbox("Potentiel *", [""] + potentials)
            pricelist = st.selectbox("Liste de prix", [""] + pricelists)
        with c2:
            company_registry = st.text_input("Registre de Commerce (RC)")
            vat = st.text_input("NIF (N° Identification Fiscale)")
            tax_article = st.text_input("Article d'Imposition")
            sin = st.text_input("NIS (N° Identification Statistique)")

        submitted = st.form_submit_button("Ajouter la pharmacie")

    if submitted:
        name = selected_name
        errors = []
        if not name.strip():
            errors.append("Le nom est obligatoire.")
        if not wilaya:
            errors.append("La wilaya est obligatoire.")
        if not commune:
            errors.append("La commune est obligatoire.")
        if not sector:
            errors.append("Le secteur est obligatoire.")
        if not street.strip():
            errors.append("L'adresse est obligatoire.")
        if not phone.strip():
            errors.append("Le téléphone est obligatoire.")
        if not potential:
            errors.append("Le potentiel est obligatoire.")
        delegate = selected_delegate
        if not delegate:
            errors.append("Le délégué est obligatoire (sélectionnez depuis la liste).")

        # Check duplicates in session AND in existing Excel file
        if name.strip() and commune:
            dup = st.session_state.pharmacies[
                (st.session_state.pharmacies["name"].str.upper() == name.strip().upper())
                & (st.session_state.pharmacies["Commune"].str.upper() == commune.upper())
            ]
            if not dup.empty:
                errors.append("Ce compte existe déjà dans la session (même nom + commune).")
            existing_xl = load_existing_records(PHARMACIES_FILE, "Comptes")
            if not existing_xl.empty and "name" in existing_xl.columns and "Commune" in existing_xl.columns:
                dup_xl = existing_xl[
                    (existing_xl["name"].astype(str).str.upper() == name.strip().upper())
                    & (existing_xl["Commune"].astype(str).str.upper() == commune.upper())
                ]
                if not dup_xl.empty:
                    errors.append("Ce compte existe déjà dans le fichier Excel (même nom + commune).")

        if errors:
            for e in errors:
                st.error(e)
        else:
            row = {
                "name": name.strip(),
                "ref": "",
                "type_id": type_id,
                "email": email.strip(),
                "phone": phone.strip(),
                "mobile": mobile.strip(),
                "customer_potential_id/id": potential or "",
                "street": street.strip(),
                "street2": "",
                "country_id": "Algérie",
                "state_id": wilaya,
                "Commune": commune,
                "city_id/id": get_city_id(communes_df, commune, wilaya),
                "zip": "",
                "Secteur (nom complet)": sector,
                "sector_id/id": get_sector_id(sectors_df, sector),
                "company_registry": company_registry.strip(),
                "vat": vat.strip(),
                "tax_article": tax_article.strip(),
                "sin": sin.strip(),
                "property_product_pricelist": pricelist or "",
                "static_portfolio_user_ids": delegate,
            }
            row_df = pd.DataFrame([row])
            try:
                r_start, r_end = append_to_excel(
                    PHARMACIES_FILE, "Comptes", PHARMACY_COLUMNS, row_df,
                )
                st.success(f"✅ Compte **{row['name']}** enregistré dans Pharmacies.xlsx (ligne {r_start}) !")
            except Exception as exc:
                st.error(f"❌ Erreur d'écriture : {exc}")
            st.session_state.pharmacies = pd.concat(
                [st.session_state.pharmacies, row_df], ignore_index=True
            )

    # --- Preview ---
    if not st.session_state.pharmacies.empty:
        st.subheader(f"📋 Comptes ajoutés cette session ({len(st.session_state.pharmacies)})")
        st.dataframe(
            st.session_state.pharmacies[
                ["name", "type_id", "state_id", "Commune",
                 "Secteur (nom complet)", "phone", "static_portfolio_user_ids"]
            ],
            use_container_width=True, hide_index=True,
        )

        col_a, col_b = st.columns(2)
        with col_a:
            buf = generate_excel_download(st.session_state.pharmacies, PHARMACY_COLUMNS, "Comptes")
            st.download_button(
                "📥 Télécharger une copie", buf,
                file_name="Pharmacies_import.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col_b:
            if st.button("🗑️ Réinitialiser pharmacies"):
                st.session_state.pharmacies = pd.DataFrame(columns=PHARMACY_COLUMNS)
                st.rerun()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    st.set_page_config(page_title="Magpharm - Création CRM", page_icon="💊", layout="wide")

    # --- Custom CSS (charte graphique Magpharm : gris #6C6E6F + rouge #E4211A) ---
    st.markdown("""
    <style>
        /* Header bar */
        .main-header {
            background: #FFFFFF;
            padding: 1.2rem 2rem;
            border-radius: 10px;
            margin-bottom: 1.5rem;
            display: flex;
            align-items: center;
            gap: 1.5rem;
            border-bottom: 3px solid #E4211A;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        }
        .main-header img {
            height: 60px;
        }
        .main-header h1 {
            color: #6C6E6F;
            font-size: 1.8rem;
            margin: 0;
        }
        .main-header p {
            color: #999999;
            font-size: 0.95rem;
            margin: 0;
        }
        /* Sidebar styling */
        [data-testid="stSidebar"] {
            background-color: #F5F5F5;
        }
        /* Buttons */
        .stButton > button[kind="primary"],
        .stFormSubmitButton > button {
            background-color: #E4211A !important;
            color: white !important;
            border: none !important;
        }
        .stButton > button[kind="primary"]:hover,
        .stFormSubmitButton > button:hover {
            background-color: #C41B15 !important;
        }
        /* Metrics */
        [data-testid="stMetricValue"] {
            color: #E4211A;
        }
        /* Subheaders */
        .stSubheader, h2, h3 {
            color: #6C6E6F !important;
        }
        /* Section dividers */
        hr {
            border-color: #E4211A !important;
        }
    </style>
    """, unsafe_allow_html=True)

    # --- Logo + Title header ---
    import base64
    logo_path = os.path.join(BASE_DIR, "Magpharm-logo.jpeg")
    with open(logo_path, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
    st.markdown(f"""
    <div class="main-header">
        <img src="data:image/jpeg;base64,{logo_b64}" alt="Magpharm">
        <div>
            <h1>Création CRM Médecins & Pharmacies</h1>
            <p>Plateforme de saisie pour importation Odoo</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    init_state()

    # --- Sidebar with logo ---
    st.sidebar.markdown(
        f'<div style="text-align:center;padding:0.5rem 0;">'
        f'<img src="data:image/png;base64,{logo_b64}" style="width:180px;">'
        f'</div>',
        unsafe_allow_html=True
    )
    st.sidebar.markdown("---")
    menu = st.sidebar.radio("Navigation", ["Créer un Médecin", "Créer une Pharmacie"])
    st.sidebar.markdown("---")
    st.sidebar.metric("Mes médecins (session)", len(st.session_state.doctors))
    st.sidebar.metric("Mes pharmacies (session)", len(st.session_state.pharmacies))

    # --- Admin section (password-protected) ---
    st.sidebar.markdown("---")
    st.sidebar.subheader("🔒 Espace Admin")
    admin_pwd = st.sidebar.text_input("Mot de passe", type="password", key="admin_pwd")
    if admin_pwd == "magpharm2026":
        existing_docs = load_existing_records(MEDECINS_FILE, "Contacts")
        existing_pha = load_existing_records(PHARMACIES_FILE, "Comptes")
        n_docs = len(existing_docs.dropna(subset=["name"])) if not existing_docs.empty and "name" in existing_docs.columns else 0
        n_pha = len(existing_pha.dropna(subset=["name"])) if not existing_pha.empty and "name" in existing_pha.columns else 0
        st.sidebar.caption(f"📂 Total : {n_docs} médecins / {n_pha} pharmacies")

        st.sidebar.subheader("📥 Télécharger les fichiers")
        with open(MEDECINS_FILE, "rb") as f:
            st.sidebar.download_button(
                f"Médecins.xlsx ({n_docs} fiches)",
                f.read(),
                file_name="Médecins.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with open(PHARMACIES_FILE, "rb") as f:
            st.sidebar.download_button(
                f"Pharmacies.xlsx ({n_pha} fiches)",
                f.read(),
                file_name="Pharmacies.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        st.sidebar.subheader("🗑️ Réinitialiser les données")
        st.sidebar.caption("Remet les fichiers Excel à zéro (vide les fiches enregistrées).")
        col_r1, col_r2 = st.sidebar.columns(2)
        with col_r1:
            if st.button("Vider Médecins", key="reset_med"):
                shutil.copy2(MEDECINS_TEMPLATE, MEDECINS_FILE)
                st.session_state.doctors = pd.DataFrame(columns=DOCTOR_COLUMNS)
                st.sidebar.success("Médecins.xlsx vidé !")
                st.rerun()
        with col_r2:
            if st.button("Vider Pharmacies", key="reset_pha"):
                shutil.copy2(PHARMACIES_TEMPLATE, PHARMACIES_FILE)
                st.session_state.pharmacies = pd.DataFrame(columns=PHARMACY_COLUMNS)
                st.sidebar.success("Pharmacies.xlsx vidé !")
                st.rerun()
    elif admin_pwd:
        st.sidebar.error("Mot de passe incorrect")

    if menu == "Créer un Médecin":
        doctor_form()
    else:
        pharmacy_form()


if __name__ == "__main__":
    main()
